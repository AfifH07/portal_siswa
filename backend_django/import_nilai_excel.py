"""
Script untuk mengimport data nilai dari file Excel LEGER ke database.
Jalankan: python import_nilai_excel.py

Format Excel LEGER:
- Row 1-3: Header sekolah
- Row 4: Header kolom (NO, NAMA SISWA, NISN, NIS, MATA PELAJARAN)
- Row 5: Nama mata pelajaran (repeat setiap ~7 kolom)
- Row 7: Label semester (Smt1, Smt2, ..., Smt6, rerata)
- Row 8+: Data siswa
"""

import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend_django.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
from apps.grades.models import Grade
from apps.students.models import Student


def parse_leger_structure(sheet):
    """
    Parse struktur file LEGER untuk mendapatkan mapping kolom.
    Returns: (nisn_col, nama_col, mapel_structure)
    mapel_structure: list of (col_smt6, mapel_name)
    """
    # Kolom tetap
    nisn_col = 3  # Column C
    nama_col = 2  # Column B

    # Parse nama mata pelajaran dari Row 5 dan mapping ke kolom Smt6
    mapel_structure = []

    # Row 5 berisi nama mapel
    row5_values = [sheet.cell(row=5, column=c).value for c in range(1, sheet.max_column + 1)]

    # Row 7 berisi label semester
    row7_values = [sheet.cell(row=7, column=c).value for c in range(1, sheet.max_column + 1)]

    current_mapel = None

    for col_idx in range(5, len(row5_values) + 1):  # Mulai dari kolom E (5)
        # Cek apakah ada nama mapel baru di row 5
        mapel_val = row5_values[col_idx - 1] if col_idx - 1 < len(row5_values) else None
        if mapel_val and str(mapel_val).strip():
            current_mapel = str(mapel_val).strip()

        # Cek apakah kolom ini adalah Smt6
        smt_val = row7_values[col_idx - 1] if col_idx - 1 < len(row7_values) else None
        if smt_val and str(smt_val).strip().lower() == 'smt6':
            if current_mapel:
                mapel_structure.append((col_idx, current_mapel))

    return nisn_col, nama_col, mapel_structure


def import_grades_from_leger(filepath, kelas, semester='Genap', tahun_ajaran='2024/2025', guru='Admin'):
    """Import data nilai dari file LEGER Excel."""
    print(f"\n{'='*60}")
    print(f"Importing: {os.path.basename(filepath)}")
    print(f"Kelas: {kelas}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Parse struktur
    nisn_col, nama_col, mapel_structure = parse_leger_structure(sheet)

    print(f"\nStruktur terdeteksi:")
    print(f"  NISN column: {nisn_col}")
    print(f"  Mata Pelajaran ({len(mapel_structure)} mapel):")
    for col, mapel in mapel_structure:
        print(f"    - Col {col}: {mapel}")

    # Import data mulai row 8
    success_count = 0
    error_count = 0
    student_count = 0

    for row_idx in range(8, sheet.max_row + 1):
        nisn_value = sheet.cell(row=row_idx, column=nisn_col).value
        nama_value = sheet.cell(row=row_idx, column=nama_col).value

        if not nisn_value:
            continue

        nisn_str = str(nisn_value).strip()

        # Skip jika bukan NISN valid
        if not nisn_str.isdigit() or len(nisn_str) < 5:
            continue

        student_count += 1

        # Cari student di database
        try:
            student = Student.objects.get(nisn=nisn_str)
        except Student.DoesNotExist:
            print(f"  WARNING: NISN {nisn_str} ({nama_value}) tidak ditemukan di database")
            error_count += 1
            continue

        # Import nilai untuk setiap mata pelajaran
        for col_idx, mapel_name in mapel_structure:
            nilai_value = sheet.cell(row=row_idx, column=col_idx).value

            if nilai_value is None or str(nilai_value).strip() == '':
                continue

            try:
                nilai = int(float(nilai_value))

                # Validasi nilai
                if nilai < 0 or nilai > 100:
                    continue

                # Cek existing atau create
                existing = Grade.objects.filter(
                    nisn=student,
                    mata_pelajaran=mapel_name,
                    semester=semester,
                    tahun_ajaran=tahun_ajaran,
                    jenis='UAS'
                ).first()

                if existing:
                    existing.nilai = nilai
                    existing.kelas = kelas
                    existing.guru = guru
                    existing.save()
                else:
                    Grade.objects.create(
                        nisn=student,
                        mata_pelajaran=mapel_name,
                        nilai=nilai,
                        semester=semester,
                        tahun_ajaran=tahun_ajaran,
                        jenis='UAS',
                        kelas=kelas,
                        guru=guru
                    )

                success_count += 1

            except (ValueError, TypeError) as e:
                error_count += 1

    wb.close()

    print(f"\nHasil Import untuk {kelas}:")
    print(f"  Siswa ditemukan: {student_count}")
    print(f"  Nilai berhasil: {success_count}")
    print(f"  Error: {error_count}")

    return success_count


def main():
    # Path ke folder data nilai
    data_folder = r'C:\Users\Afif H\.vscode\belajaroiii\semester 6\portal-siswa\data nilai'

    print("="*80)
    print("IMPORT NILAI DARI FILE LEGER EXCEL")
    print("="*80)

    # List files
    files = [
        ('f_legersemua_XII-A (2).xlsx', 'XII A'),
        ('f_legersemua_XII-B (2).xlsx', 'XII B'),
        ('f_legersemua_XII-C (2).xlsx', 'XII C'),
    ]

    # Check files exist
    for filename, kelas in files:
        filepath = os.path.join(data_folder, filename)
        if os.path.exists(filepath):
            print(f"  [OK] {filename} -> {kelas}")
        else:
            print(f"  [MISSING] {filename}")

    print("\nMulai import data...")

    # Import
    total_success = 0

    for filename, kelas in files:
        filepath = os.path.join(data_folder, filename)
        if os.path.exists(filepath):
            count = import_grades_from_leger(
                filepath,
                kelas=kelas,
                semester='Genap',
                tahun_ajaran='2024/2025',
                guru='Admin'
            )
            total_success += count

    print("\n" + "="*80)
    print(f"TOTAL NILAI BERHASIL DIIMPORT: {total_success}")
    print("="*80)


if __name__ == '__main__':
    main()
