from rest_framework import viewsets, status, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.db.models import Avg, Max, Min, Count, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO

from apps.accounts.models import Assignment
from apps.core.models import TahunAjaran, MasterMapel


class GradePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

from .models import Grade
from .serializers import (
    GradeSerializer, GradeCreateSerializer, GradeStatsSerializer,
    GradeAverageSerializer, ClassGradesSerializer
)
from apps.accounts.permissions import IsSuperAdmin, IsPimpinan, IsGuru
from apps.students.models import Student


class GradeViewSet(viewsets.ModelViewSet):
    queryset = Grade.objects.all()
    permission_classes = [IsAuthenticated]
    pagination_class = GradePagination

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return GradeCreateSerializer
        return GradeSerializer

    def get_permissions(self):
        if self.action in ['create']:
            permission_classes = [IsAuthenticated, IsGuru | IsSuperAdmin]
        elif self.action in ['update', 'partial_update', 'destroy']:
            permission_classes = [IsAuthenticated, IsGuru | IsSuperAdmin]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def get_queryset(self):
        queryset = Grade.objects.select_related('nisn')
        user = self.request.user

        if user.role == 'walisantri':
            # Filter by linked student's NISN (linked_student_nisn is a string, nisn is FK)
            queryset = queryset.filter(nisn__nisn=user.linked_student_nisn)
        elif user.role == 'guru':
            queryset = queryset.filter(guru=user.name if hasattr(user, 'name') else user.username)
        elif user.role == 'pimpinan':
            pass
        elif user.role == 'superadmin':
            pass

        kelas = self.request.query_params.get('kelas')
        semester = self.request.query_params.get('semester')
        tahun_ajaran = self.request.query_params.get('tahun_ajaran')
        jenis = self.request.query_params.get('jenis')
        mata_pelajaran = self.request.query_params.get('mata_pelajaran')
        search = self.request.query_params.get('search')

        if kelas:
            queryset = queryset.filter(kelas=kelas)
        if semester:
            queryset = queryset.filter(semester=semester)
        if tahun_ajaran:
            queryset = queryset.filter(tahun_ajaran=tahun_ajaran)
        if jenis:
            queryset = queryset.filter(jenis=jenis)
        if mata_pelajaran:
            queryset = queryset.filter(mata_pelajaran__icontains=mata_pelajaran)
        if search:
            queryset = queryset.filter(
                Q(mata_pelajaran__icontains=search) |
                Q(nisn__nama__icontains=search)
            )

        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        guru_name = self.request.user.name if hasattr(self.request.user, 'name') and self.request.user.name else self.request.user.username
        serializer.save(guru=guru_name)

    def perform_update(self, serializer):
        instance = self.get_object()
        user_role = getattr(self.request.user, 'role', None)
        user_name = self.request.user.name if hasattr(self.request.user, 'name') and self.request.user.name else self.request.user.username

        # Superadmin can update any grade
        if user_role == 'superadmin':
            serializer.save()
        # Guru can only update their own grades
        elif instance.guru == user_name:
            serializer.save()
        else:
            return Response(
                {'success': False, 'message': 'Anda tidak memiliki izin untuk mengedit nilai ini'},
                status=status.HTTP_403_FORBIDDEN
            )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_average_grade(request, nisn):
    """
    Get average grade for a student.
    Walisantri can only access their linked student's data.
    """
    try:
        student = Student.objects.get(nisn=nisn)
    except Student.DoesNotExist:
        return Response(
            {'success': False, 'message': 'Siswa tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    user = request.user

    # Walisantri can only see their linked student's grades
    if user.role == 'walisantri':
        if user.linked_student_nisn != nisn:
            return Response(
                {'success': False, 'message': 'Anda tidak memiliki akses ke data ini'},
                status=status.HTTP_403_FORBIDDEN
            )

    semester = request.query_params.get('semester')
    tahun_ajaran = request.query_params.get('tahun_ajaran')

    grades = Grade.objects.filter(nisn=student)

    if semester:
        grades = grades.filter(semester=semester)
    if tahun_ajaran:
        grades = grades.filter(tahun_ajaran=tahun_ajaran)

    if not grades.exists():
        return Response({
            'success': True,
            'nisn': nisn,
            'nama': student.nama,
            'semester': semester or 'Semua',
            'tahun_ajaran': tahun_ajaran or 'Semua',
            'rata_rata': 0,
            'jumlah_mata_pelajaran': 0,
            'mata_pelajaran': []
        })

    mata_pelajaran_data = {}
    for grade in grades:
        mp = grade.mata_pelajaran
        if mp not in mata_pelajaran_data:
            mata_pelajaran_data[mp] = {
                'nama': mp,
                'detail': []
            }
        mata_pelajaran_data[mp]['detail'].append({
            'jenis': grade.jenis,
            'nilai': grade.nilai,
            'kelas': grade.kelas,
            'semester': grade.semester,
            'tahun_ajaran': grade.tahun_ajaran
        })

    mata_pelajaran_list = []
    total_nilai = 0
    count = 0

    for mp, data in mata_pelajaran_data.items():
        rata_rata_mp = sum(d['nilai'] for d in data['detail']) / len(data['detail'])
        mata_pelajaran_list.append({
            'nama': mp,
            'rata_rata': round(rata_rata_mp, 2),
            'detail': data['detail']
        })
        total_nilai += sum(d['nilai'] for d in data['detail'])
        count += len(data['detail'])

    rata_rata = round(total_nilai / count, 2) if count > 0 else 0

    return Response({
        'success': True,
        'nisn': nisn,
        'nama': student.nama,
        'semester': semester or 'Semua',
        'tahun_ajaran': tahun_ajaran or 'Semua',
        'rata_rata': rata_rata,
        'jumlah_mata_pelajaran': len(mata_pelajaran_list),
        'mata_pelajaran': mata_pelajaran_list
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_child_grades(request):
    """
    Get grades for walisantri's linked child.
    Uses Django ORM aggregation for optimal performance.

    Response:
    {
        "success": true,
        "nisn": "...",
        "nama": "...",
        "kelas": "...",
        "rata_rata_total": 85.5,
        "jumlah_mata_pelajaran": 6,
        "grades": [
            {
                "mata_pelajaran": "Bahasa Arab",
                "rata_rata": 87.5,
                "nilai_tertinggi": 92,
                "nilai_terendah": 80,
                "jumlah_nilai": 4
            }
        ]
    }
    """
    user = request.user

    # Only walisantri can access this endpoint
    if user.role != 'walisantri':
        return Response(
            {'success': False, 'message': 'Endpoint ini hanya untuk walisantri'},
            status=status.HTTP_403_FORBIDDEN
        )

    # Get NISN from logged-in walisantri
    nisn = getattr(user, 'linked_student_nisn', None)
    if not nisn:
        return Response(
            {'success': False, 'message': 'Akun belum terhubung dengan data siswa'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Verify student exists
    try:
        student = Student.objects.get(nisn=nisn)
    except Student.DoesNotExist:
        return Response(
            {'success': False, 'message': 'Data siswa tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Get filter params
    semester = request.query_params.get('semester')
    tahun_ajaran = request.query_params.get('tahun_ajaran')

    # Build base queryset - FILTER BY Student object (FK points to Student PK, not NISN string)
    grades = Grade.objects.filter(nisn=student)

    if semester:
        grades = grades.filter(semester=semester)
    if tahun_ajaran:
        grades = grades.filter(tahun_ajaran=tahun_ajaran)

    # Check if any grades exist
    if not grades.exists():
        return Response({
            'success': True,
            'nisn': nisn,
            'nama': student.nama,
            'kelas': student.kelas,
            'rata_rata_total': 0,
            'jumlah_mata_pelajaran': 0,
            'grades': []
        })

    # OPTIMIZED: Use Django ORM aggregation (single query)
    stats = grades.values('mata_pelajaran').annotate(
        rata_rata=Avg('nilai'),
        nilai_tertinggi=Max('nilai'),
        nilai_terendah=Min('nilai'),
        jumlah_nilai=Count('id')
    ).order_by('mata_pelajaran')

    # Calculate total average (single query)
    total_avg = grades.aggregate(avg=Avg('nilai'))['avg'] or 0

    # Convert to list and round rata_rata
    grades_list = []
    for item in stats:
        grades_list.append({
            'mata_pelajaran': item['mata_pelajaran'],
            'rata_rata': round(item['rata_rata'], 2) if item['rata_rata'] else 0,
            'nilai_tertinggi': item['nilai_tertinggi'] or 0,
            'nilai_terendah': item['nilai_terendah'] or 0,
            'jumlah_nilai': item['jumlah_nilai'] or 0
        })

    return Response({
        'success': True,
        'nisn': nisn,
        'nama': student.nama,
        'kelas': student.kelas,
        'rata_rata_total': round(total_avg, 2),
        'jumlah_mata_pelajaran': len(grades_list),
        'grades': grades_list
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsGuru | IsPimpinan | IsSuperAdmin])
def get_class_grades(request, kelas):
    user = request.user
    if user.role == 'guru' and user.kelas != kelas:
        return Response(
            {'success': False, 'message': 'Anda tidak memiliki akses ke kelas ini'},
            status=status.HTTP_403_FORBIDDEN
        )

    semester = request.query_params.get('semester')
    tahun_ajaran = request.query_params.get('tahun_ajaran')
    mata_pelajaran = request.query_params.get('mata_pelajaran')

    # Use select_related to prevent N+1 queries when accessing nisn FK
    grades = Grade.objects.select_related('nisn').filter(kelas=kelas)

    if semester:
        grades = grades.filter(semester=semester)
    if tahun_ajaran:
        grades = grades.filter(tahun_ajaran=tahun_ajaran)
    if mata_pelajaran:
        grades = grades.filter(mata_pelajaran__icontains=mata_pelajaran)

    # Optimized: Use annotate + values to aggregate in single query
    aggregated = grades.values(
        'nisn__nisn', 'nisn__nama'
    ).annotate(
        total_nilai=Sum('nilai'),
        count=Count('id')
    )

    result = []
    for data in aggregated:
        rata_rata = round(data['total_nilai'] / data['count'], 2) if data['count'] > 0 else 0
        result.append({
            'nisn': data['nisn__nisn'],
            'nama': data['nisn__nama'],
            'rata_rata': rata_rata,
            'total_nilai': data['total_nilai'],
            'jumlah_nilai': data['count']
        })

    result.sort(key=lambda x: x['rata_rata'], reverse=True)

    return Response({
        'success': True,
        'kelas': kelas,
        'semester': semester or 'Semua',
        'tahun_ajaran': tahun_ajaran or 'Semua',
        'mata_pelajaran': mata_pelajaran or 'Semua',
        'data': result
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsPimpinan | IsSuperAdmin])
def get_all_grades(request):
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 25))

    kelas = request.query_params.get('kelas')
    semester = request.query_params.get('semester')
    tahun_ajaran = request.query_params.get('tahun_ajaran')

    queryset = Grade.objects.select_related('nisn')

    if kelas:
        queryset = queryset.filter(kelas=kelas)
    if semester:
        queryset = queryset.filter(semester=semester)
    if tahun_ajaran:
        queryset = queryset.filter(tahun_ajaran=tahun_ajaran)

    queryset = queryset.order_by('-created_at')

    start = (page - 1) * page_size
    end = start + page_size

    total = queryset.count()
    grades = queryset[start:end]

    serializer = GradeSerializer(grades, many=True)

    return Response({
        'success': True,
        'count': total,
        'page': page,
        'page_size': page_size,
        'next': page * page_size < total,
        'previous': page > 1,
        'results': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_classes(request):
    classes = Student.objects.values_list('kelas', flat=True).distinct()
    return Response({
        'success': True,
        'classes': sorted(set([c for c in classes if c]))
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mata_pelajaran(request):
    mata_pelajaran = Grade.objects.values_list('mata_pelajaran', flat=True).distinct()
    return Response({
        'success': True,
        'mata_pelajaran': sorted(list(set(mata_pelajaran)))
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsGuru | IsSuperAdmin])
def import_excel_grades(request):
    """Import grades from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'File tidak ditemukan'
            }, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Format file harus .xlsx atau .xls'
            }, status=status.HTTP_400_BAD_REQUEST)

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        if len(rows) == 0:
            return Response({
                'success': False,
                'message': 'File kosong atau tidak ada data'
            }, status=status.HTTP_400_BAD_REQUEST)

        kelas = request.data.get('kelas', '')
        mata_pelajaran = request.data.get('mata_pelajaran', '')
        semester = request.data.get('semester', 'Ganjil')
        tahun_ajaran = request.data.get('tahun_ajaran', '2024/2025')
        jenis = request.data.get('jenis', 'UH')
        guru_name = request.user.name if hasattr(request.user, 'name') else request.user.username

        if not kelas:
            return Response({
                'success': False,
                'message': 'Kelas harus diisi'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not mata_pelajaran:
            return Response({
                'success': False,
                'message': 'Mata pelajaran harus diisi'
            }, status=status.HTTP_400_BAD_REQUEST)

        success_count = 0
        error_count = 0
        errors = []

        for row in rows:
            try:
                if not row[0] or not row[1]:
                    error_count += 1
                    errors.append(f'Baris tidak lengkap: {row}')
                    continue

                nisn = str(row[0]).strip()
                nilai = float(row[1])

                if nilai < 0 or nilai > 100:
                    error_count += 1
                    errors.append(f'Nilai {nilai} tidak valid (harus 0-100) untuk NISN {nisn}')
                    continue

                try:
                    student = Student.objects.get(nisn=nisn)
                except Student.DoesNotExist:
                    error_count += 1
                    errors.append(f'Siswa dengan NISN {nisn} tidak ditemukan')
                    continue

                if student.kelas != kelas:
                    error_count += 1
                    errors.append(f'Siswa {nisn} bukan dari kelas {kelas} (dari {student.kelas})')
                    continue

                existing_grade = Grade.objects.filter(
                    nisn=student,
                    mata_pelajaran=mata_pelajaran,
                    semester=semester,
                    tahun_ajaran=tahun_ajaran,
                    jenis=jenis
                ).first()

                if existing_grade:
                    existing_grade.nilai = int(nilai)
                    existing_grade.guru = guru_name
                    existing_grade.save()
                else:
                    Grade.objects.create(
                        nisn=student,
                        mata_pelajaran=mata_pelajaran,
                        nilai=int(nilai),
                        semester=semester,
                        tahun_ajaran=tahun_ajaran,
                        jenis=jenis,
                        kelas=kelas,
                        guru=guru_name
                    )

                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f'Error memproses baris: {str(e)}')

        return Response({
            'success': True,
            'message': f'Import selesai: {success_count} berhasil, {error_count} gagal',
            'total_rows': len(rows),
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors if errors else None
        })

    except Exception as e:
        return Response({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_statistics(request):
    """
    Get grade statistics for admin/guru dashboard.
    Returns: rata-rata kelas, % ketuntasan, siswa di bawah rata-rata, total records,
    distribution data for charts, trend data, top students, and lowest subjects.
    """
    user = request.user
    KKM = 75  # Kriteria Ketuntasan Minimal

    # Build base queryset based on user role
    queryset = Grade.objects.select_related('nisn')

    if user.role == 'guru':
        queryset = queryset.filter(guru=user.name if hasattr(user, 'name') else user.username)
    elif user.role == 'walisantri':
        queryset = queryset.filter(nisn__nisn=user.linked_student_nisn)

    # Apply filters from query params
    search = request.query_params.get('search')
    kelas = request.query_params.get('kelas')
    semester = request.query_params.get('semester')
    tahun_ajaran = request.query_params.get('tahun_ajaran')
    mata_pelajaran = request.query_params.get('mata_pelajaran')
    jenis = request.query_params.get('jenis')

    if search:
        queryset = queryset.filter(
            Q(mata_pelajaran__icontains=search) |
            Q(nisn__nama__icontains=search) |
            Q(nisn__nisn__icontains=search)
        )
    if kelas:
        queryset = queryset.filter(kelas=kelas)
    if semester:
        queryset = queryset.filter(semester=semester)
    if tahun_ajaran:
        queryset = queryset.filter(tahun_ajaran=tahun_ajaran)
    if mata_pelajaran:
        queryset = queryset.filter(mata_pelajaran__icontains=mata_pelajaran)
    if jenis:
        queryset = queryset.filter(jenis=jenis)

    # Basic statistics
    stats = queryset.aggregate(
        total_grades=Count('id'),
        average_score=Coalesce(Avg('nilai'), 0.0),
        max_score=Max('nilai'),
        min_score=Min('nilai')
    )

    total_grades = stats['total_grades'] or 0
    average_score = round(stats['average_score'], 2) if stats['average_score'] else 0
    max_score = stats['max_score'] or 0
    min_score = stats['min_score'] or 0

    # Ketuntasan calculation
    above_kkm = queryset.filter(nilai__gte=KKM).count()
    below_kkm = queryset.filter(nilai__lt=KKM).count()
    ketuntasan_percentage = round((above_kkm / total_grades * 100), 1) if total_grades > 0 else 0

    # ========== NEW: Unique students needing remedial ==========
    # Count unique students with at least one grade below KKM
    unique_remedial_students = queryset.filter(nilai__lt=KKM).values('nisn').distinct().count()

    # ========== NEW: Total unique subjects ==========
    total_subjects = queryset.values('mata_pelajaran').distinct().count()

    # ========== NEW: Top 3 Students with Highest Average ==========
    top_students_data = queryset.values(
        'nisn__nisn', 'nisn__nama', 'nisn__kelas'
    ).annotate(
        avg_nilai=Avg('nilai'),
        total_grades=Count('id')
    ).order_by('-avg_nilai')[:3]

    top_students = []
    for idx, student in enumerate(top_students_data, 1):
        top_students.append({
            'rank': idx,
            'nisn': student['nisn__nisn'],
            'nama': student['nisn__nama'],
            'kelas': student['nisn__kelas'],
            'rata_rata': round(student['avg_nilai'], 1) if student['avg_nilai'] else 0,
            'jumlah_nilai': student['total_grades']
        })

    # ========== NEW: Lowest 5 Subjects by Average ==========
    lowest_subjects_data = queryset.values('mata_pelajaran').annotate(
        avg_nilai=Avg('nilai'),
        total_grades=Count('id'),
        below_kkm_count=Count('id', filter=Q(nilai__lt=KKM))
    ).order_by('avg_nilai')[:5]

    lowest_subjects = []
    for subject in lowest_subjects_data:
        if subject['mata_pelajaran']:
            lowest_subjects.append({
                'nama': subject['mata_pelajaran'],
                'rata_rata': round(subject['avg_nilai'], 1) if subject['avg_nilai'] else 0,
                'total_nilai': subject['total_grades'],
                'perlu_perbaikan': subject['below_kkm_count']
            })

    # ========== IMPROVED: Grade distribution counting UNIQUE STUDENTS per range ==========
    # Count unique students in each grade range (by their average score)
    student_averages = queryset.values('nisn').annotate(
        avg_nilai=Avg('nilai')
    )

    distribution = {
        '0-50': 0,
        '51-60': 0,
        '61-70': 0,
        '71-80': 0,
        '81-90': 0,
        '91-100': 0
    }

    for student in student_averages:
        avg = student['avg_nilai'] or 0
        if avg <= 50:
            distribution['0-50'] += 1
        elif avg <= 60:
            distribution['51-60'] += 1
        elif avg <= 70:
            distribution['61-70'] += 1
        elif avg <= 80:
            distribution['71-80'] += 1
        elif avg <= 90:
            distribution['81-90'] += 1
        else:
            distribution['91-100'] += 1

    # Trend data by jenis penilaian - with fixed order
    # Order: Penugasan → Tes Tulis → Tes Lisan → Portofolio → Praktek → Proyek → UTS → UAS → Lainnya
    JENIS_ORDER = [
        'penugasan', 'tes_tulis', 'tes_lisan', 'portofolio',
        'praktek', 'proyek', 'uts', 'uas'
    ]
    JENIS_LABELS = {
        'penugasan': 'Penugasan',
        'tes_tulis': 'Tes Tulis',
        'tes_lisan': 'Tes Lisan',
        'portofolio': 'Portofolio',
        'praktek': 'Praktek',
        'proyek': 'Proyek',
        'uts': 'UTS',
        'uas': 'UAS',
        # Legacy values
        'UH': 'Lainnya',
        'UTS': 'UTS (Lama)',
        'UAS': 'UAS (Lama)',
        'Tugas': 'Lainnya',
        'Proyek': 'Lainnya',
    }

    # Get all jenis with their averages
    jenis_trend_raw = queryset.values('jenis').annotate(
        avg_nilai=Avg('nilai'),
        count=Count('id')
    )

    # Build a dict for quick lookup
    jenis_avg_map = {}
    legacy_values = []
    for item in jenis_trend_raw:
        if item['jenis']:
            jenis_val = item['jenis']
            avg_val = round(item['avg_nilai'], 2) if item['avg_nilai'] else 0
            if jenis_val in JENIS_ORDER:
                jenis_avg_map[jenis_val] = avg_val
            else:
                # Legacy values - collect for "Lainnya"
                legacy_values.append(avg_val)

    # Build trend arrays in fixed order
    trend_labels = []
    trend_data = []
    for jenis in JENIS_ORDER:
        if jenis in jenis_avg_map:
            trend_labels.append(JENIS_LABELS.get(jenis, jenis))
            trend_data.append(jenis_avg_map[jenis])

    # Add "Lainnya" for legacy values if any exist
    if legacy_values:
        avg_lainnya = round(sum(legacy_values) / len(legacy_values), 2)
        trend_labels.append('Lainnya')
        trend_data.append(avg_lainnya)

    # Ketuntasan data for doughnut chart
    ketuntasan_data = {
        'tuntas': above_kkm,
        'tidak_tuntas': below_kkm
    }

    # Class comparison (for admin/pimpinan)
    class_comparison = []
    if user.role in ['superadmin', 'pimpinan']:
        class_stats = queryset.values('kelas').annotate(
            avg_nilai=Avg('nilai'),
            count=Count('id'),
            tuntas=Count('id', filter=Q(nilai__gte=KKM))
        ).order_by('kelas')

        for item in class_stats:
            if item['kelas']:
                class_comparison.append({
                    'kelas': item['kelas'],
                    'rata_rata': round(item['avg_nilai'], 2) if item['avg_nilai'] else 0,
                    'total': item['count'],
                    'tuntas': item['tuntas'],
                    'persen_tuntas': round((item['tuntas'] / item['count'] * 100), 1) if item['count'] > 0 else 0
                })

    return Response({
        'success': True,
        'statistics': {
            'total_grades': total_grades,
            'average_score': average_score,
            'max_score': max_score,
            'min_score': min_score,
            'above_average': above_kkm,
            'below_average': below_kkm,
            'ketuntasan_percentage': ketuntasan_percentage,
            'unique_remedial_students': unique_remedial_students,
            'total_subjects': total_subjects,
            'kkm': KKM
        },
        'distribution': distribution,
        'trend': {
            'labels': trend_labels,
            'data': trend_data
        },
        'ketuntasan': ketuntasan_data,
        'top_students': top_students,
        'lowest_subjects': lowest_subjects,
        'class_comparison': class_comparison
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsGuru | IsSuperAdmin])
def generate_template(request):
    """
    Generate Emerald-styled Excel template for grade input.
    Features:
    - Emerald header (#178560) with white bold text
    - Instruction row with light emerald background
    - Protected NISN & Nama columns (read-only)
    - Data validation for Nilai (0-100)

    Required params: kelas, mata_pelajaran
    Optional params: semester, tahun_ajaran, jenis
    """
    try:
        # Debug: Log incoming request
        print(f"[generate_template] Request params: {dict(request.query_params)}")

        # Get parameters from query string with safe defaults
        kelas_raw = request.query_params.get('kelas', None)
        mata_pelajaran_raw = request.query_params.get('mata_pelajaran', None)

        # Handle None values explicitly
        kelas = kelas_raw.strip() if kelas_raw else ''
        mata_pelajaran = mata_pelajaran_raw.strip() if mata_pelajaran_raw else ''
        semester = (request.query_params.get('semester') or 'Ganjil').strip()
        tahun_ajaran = (request.query_params.get('tahun_ajaran') or '2024/2025').strip()
        jenis = (request.query_params.get('jenis') or 'UH').strip()

        print(f"[generate_template] Parsed: kelas='{kelas}', mapel='{mata_pelajaran}', jenis='{jenis}'")

        # Validate required parameters
        if not kelas:
            return Response({
                'success': False,
                'message': 'Parameter kelas wajib diisi',
                'received': {'kelas': kelas_raw}
            }, status=status.HTTP_400_BAD_REQUEST)

        if not mata_pelajaran:
            return Response({
                'success': False,
                'message': 'Parameter mata_pelajaran wajib diisi',
                'received': {'mata_pelajaran': mata_pelajaran_raw}
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get students in the specified class
        try:
            # Use exact match for kelas field
            students = Student.objects.filter(kelas=kelas, aktif=True).order_by('nama')
            student_count = students.count()
            print(f"[generate_template] Found {student_count} students in kelas '{kelas}'")
        except Exception as e:
            print(f"[generate_template] Query error: {str(e)}")
            return Response({
                'success': False,
                'message': f'Error query siswa: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if student_count == 0:
            # Try case-insensitive search as fallback
            students = Student.objects.filter(kelas__iexact=kelas, aktif=True).order_by('nama')
            student_count = students.count()
            print(f"[generate_template] Case-insensitive search found {student_count} students")

        if student_count == 0:
            return Response({
                'success': False,
                'message': f'Tidak ada siswa aktif di kelas "{kelas}"',
                'hint': 'Pastikan nama kelas sesuai dengan data di database'
            }, status=status.HTTP_404_NOT_FOUND)

        # Convert to list to ensure it's not a lazy queryset
        students_list = list(students)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Nilai"

        # ==========================================
        # EMERALD THEME COLORS
        # ==========================================
        EMERALD_DARK = "178560"      # Header background
        EMERALD_LIGHT = "D1FAE5"     # Instruction row background
        EMERALD_ACCENT = "34C99A"    # Accent color
        WHITE = "FFFFFF"
        TEXT_DARK = "1F2937"
        TEXT_MUTED = "6B7280"

        # ==========================================
        # STYLES
        # ==========================================
        # Header style (Row 1) - Emerald Dark with White Bold text
        header_font = Font(bold=True, color=WHITE, size=11, name='Calibri')
        header_fill = PatternFill(start_color=EMERALD_DARK, end_color=EMERALD_DARK, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Instruction style (Row 2) - Light Emerald with muted text
        instruction_font = Font(italic=True, color=TEXT_MUTED, size=10, name='Calibri')
        instruction_fill = PatternFill(start_color=EMERALD_LIGHT, end_color=EMERALD_LIGHT, fill_type="solid")
        instruction_alignment = Alignment(horizontal="center", vertical="center")

        # Data cell styles
        cell_font = Font(color=TEXT_DARK, size=10, name='Calibri')
        cell_alignment = Alignment(horizontal="left", vertical="center")
        cell_alignment_center = Alignment(horizontal="center", vertical="center")

        # Locked cell style (NISN, Nama - read-only)
        locked_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        locked_protection = Protection(locked=True)

        # Unlocked cell style (Nilai, Keterangan - editable)
        unlocked_protection = Protection(locked=False)
        editable_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

        # Border styles
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        header_border = Border(
            left=Side(style='medium', color=EMERALD_DARK),
            right=Side(style='medium', color=EMERALD_DARK),
            top=Side(style='medium', color=EMERALD_DARK),
            bottom=Side(style='medium', color=EMERALD_DARK)
        )

        # ==========================================
        # HEADER ROW (Row 1)
        # ==========================================
        headers = ['No', 'NISN', 'Nama Siswa', 'Nilai', 'Keterangan']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # ==========================================
        # INSTRUCTION ROW (Row 2)
        # ==========================================
        instructions = [
            '#',
            '(Jangan diubah)',
            '(Jangan diubah)',
            'Isi nilai 0-100',
            '(Opsional)'
        ]
        for col, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=2, column=col, value=instruction)
            cell.font = instruction_font
            cell.fill = instruction_fill
            cell.alignment = instruction_alignment
            cell.border = thin_border

        # ==========================================
        # COLUMN WIDTHS
        # ==========================================
        ws.column_dimensions['A'].width = 5    # No
        ws.column_dimensions['B'].width = 18   # NISN
        ws.column_dimensions['C'].width = 35   # Nama Siswa
        ws.column_dimensions['D'].width = 12   # Nilai
        ws.column_dimensions['E'].width = 25   # Keterangan

        # ==========================================
        # DATA ROWS (Starting from Row 3)
        # ==========================================
        for idx, student in enumerate(students_list, 1):
            row = idx + 2  # Start from row 3 (after header and instruction)

            # Column A: No (locked)
            no_cell = ws.cell(row=row, column=1, value=idx)
            no_cell.font = cell_font
            no_cell.alignment = cell_alignment_center
            no_cell.fill = locked_fill
            no_cell.border = thin_border
            no_cell.protection = locked_protection

            # Column B: NISN (locked, formatted as text to preserve leading zeros)
            nisn_cell = ws.cell(row=row, column=2)
            nisn_cell.value = str(student.nisn)  # Force string
            nisn_cell.number_format = '@'  # Text format
            nisn_cell.font = Font(color=TEXT_DARK, size=10, name='Calibri', bold=True)
            nisn_cell.alignment = cell_alignment
            nisn_cell.fill = locked_fill
            nisn_cell.border = thin_border
            nisn_cell.protection = locked_protection

            # Column C: Nama Siswa (locked)
            nama_cell = ws.cell(row=row, column=3, value=student.nama)
            nama_cell.font = cell_font
            nama_cell.alignment = cell_alignment
            nama_cell.fill = locked_fill
            nama_cell.border = thin_border
            nama_cell.protection = locked_protection

            # Column D: Nilai (editable - light emerald background)
            nilai_cell = ws.cell(row=row, column=4, value='')
            nilai_cell.font = Font(color=EMERALD_DARK, size=11, name='Calibri', bold=True)
            nilai_cell.alignment = cell_alignment_center
            nilai_cell.fill = editable_fill
            nilai_cell.border = thin_border
            nilai_cell.protection = unlocked_protection

            # Column E: Keterangan (editable)
            ket_cell = ws.cell(row=row, column=5, value='')
            ket_cell.font = cell_font
            ket_cell.alignment = cell_alignment
            ket_cell.fill = editable_fill
            ket_cell.border = thin_border
            ket_cell.protection = unlocked_protection

        # ==========================================
        # DATA VALIDATION (Nilai: 0-100)
        # ==========================================
        last_data_row = len(students_list) + 2
        nilai_validation = DataValidation(
            type="whole",
            operator="between",
            formula1=0,
            formula2=100,
            showErrorMessage=True,
            errorTitle="Nilai Tidak Valid",
            error="Masukkan angka antara 0 sampai 100.",
            showInputMessage=True,
            promptTitle="Input Nilai",
            prompt="Masukkan nilai siswa (0-100)"
        )
        ws.add_data_validation(nilai_validation)
        nilai_validation.add(f'D3:D{last_data_row}')

        # ==========================================
        # SHEET PROTECTION
        # ==========================================
        # Protect the sheet but allow editing unlocked cells (no password needed)
        ws.protection.sheet = True
        ws.protection.enable()
        # Note: Don't set password=None as openpyxl tries to hash it

        # ==========================================
        # METADATA ROW (Hidden info for import)
        # ==========================================
        # Add metadata in a hidden row for import reference
        metadata_row = last_data_row + 2
        ws.cell(row=metadata_row, column=1, value='[METADATA]')
        ws.cell(row=metadata_row, column=2, value=f'kelas={kelas}')
        ws.cell(row=metadata_row, column=3, value=f'mapel={mata_pelajaran}')
        ws.cell(row=metadata_row, column=4, value=f'semester={semester}')
        ws.cell(row=metadata_row, column=5, value=f'jenis={jenis}')

        # Style metadata row (light gray, small font)
        meta_font = Font(color="9CA3AF", size=8, italic=True)
        for col in range(1, 6):
            cell = ws.cell(row=metadata_row, column=col)
            cell.font = meta_font
            cell.protection = locked_protection

        # Hide metadata row
        ws.row_dimensions[metadata_row].hidden = True

        # ==========================================
        # FREEZE PANES & ROW HEIGHT
        # ==========================================
        ws.freeze_panes = 'A3'  # Freeze header and instruction rows
        ws.row_dimensions[1].height = 28  # Header row height
        ws.row_dimensions[2].height = 22  # Instruction row height

        # ==========================================
        # GENERATE FILENAME & RESPONSE
        # ==========================================
        safe_kelas = kelas.replace(' ', '_')
        safe_mapel = mata_pelajaran.replace(' ', '_')[:20]
        filename = f"Template_Nilai_{safe_kelas}_{safe_mapel}_{jenis}.xlsx"

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'

        return response

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] generate_template: {str(e)}")
        print(error_details)
        return Response({
            'success': False,
            'message': f'Gagal membuat template: {str(e)}',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsGuru | IsSuperAdmin])
def import_grades_v2(request):
    """
    Import grades from Excel file (v2 - improved version).
    Uses NISN as unique identifier and auto-assigns guru from request.user.

    Supports TWO template formats:

    SIMPLIFIED FORMAT (New Emerald Template):
    - Row 1: Header (No, NISN, Nama Siswa, Nilai, Keterangan)
    - Row 2: Instructions (skipped)
    - Row 3+: Data rows
    - Hidden metadata row: [METADATA] with kelas, mapel, semester, jenis

    LEGACY FORMAT:
    - Column A-H: No, NISN, Nama, Mapel, Semester, Tahun, Jenis, Nilai
    - Row 2+: Data rows
    """
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'File tidak ditemukan'
            }, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Format file harus .xlsx atau .xls'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get metadata from request (can override Excel values)
        default_kelas = request.data.get('kelas', '')
        default_mata_pelajaran = request.data.get('mata_pelajaran', '')
        default_semester = request.data.get('semester', 'Ganjil')
        default_tahun_ajaran = request.data.get('tahun_ajaran', '2024/2025')
        default_jenis = request.data.get('jenis', 'UH')

        # Get guru name from authenticated user
        guru_name = request.user.name if hasattr(request.user, 'name') and request.user.name else request.user.username

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # ==========================================
        # DETECT TEMPLATE FORMAT
        # ==========================================
        header_row = [cell.value for cell in sheet[1]]
        header_count = len([h for h in header_row if h])

        # Check if this is the simplified emerald template (5 columns)
        is_simplified = header_count <= 5 and any(
            str(h).lower() == 'keterangan' for h in header_row if h
        )

        # Try to read metadata from hidden row
        metadata_from_excel = {}
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == '[METADATA]':
                for cell in row[1:]:
                    if cell and '=' in str(cell):
                        key, value = str(cell).split('=', 1)
                        metadata_from_excel[key.strip()] = value.strip()
                break

        # Override defaults with Excel metadata if available
        if metadata_from_excel:
            if 'kelas' in metadata_from_excel and not default_kelas:
                default_kelas = metadata_from_excel['kelas']
            if 'mapel' in metadata_from_excel and not default_mata_pelajaran:
                default_mata_pelajaran = metadata_from_excel['mapel']
            if 'semester' in metadata_from_excel and not default_semester:
                default_semester = metadata_from_excel['semester']
            if 'jenis' in metadata_from_excel and not default_jenis:
                default_jenis = metadata_from_excel['jenis']

        # ==========================================
        # MAP COLUMNS
        # ==========================================
        col_map = {}
        for idx, header in enumerate(header_row):
            if header:
                header_lower = str(header).lower().strip()
                if 'nisn' in header_lower:
                    col_map['nisn'] = idx
                elif 'nama' in header_lower:
                    col_map['nama'] = idx
                elif 'mata' in header_lower or 'pelajaran' in header_lower or 'mapel' in header_lower:
                    col_map['mata_pelajaran'] = idx
                elif 'semester' in header_lower:
                    col_map['semester'] = idx
                elif 'tahun' in header_lower:
                    col_map['tahun_ajaran'] = idx
                elif 'jenis' in header_lower:
                    col_map['jenis'] = idx
                elif 'nilai' in header_lower:
                    col_map['nilai'] = idx
                elif 'keterangan' in header_lower:
                    col_map['keterangan'] = idx

        # Default positions based on template format
        if is_simplified:
            # Simplified template: No(0), NISN(1), Nama(2), Nilai(3), Keterangan(4)
            if 'nisn' not in col_map:
                col_map['nisn'] = 1
            if 'nilai' not in col_map:
                col_map['nilai'] = 3
        else:
            # Legacy template: No(0), NISN(1), Nama(2), Mapel(3), Semester(4), Tahun(5), Jenis(6), Nilai(7)
            if 'nisn' not in col_map:
                col_map['nisn'] = 1
            if 'nilai' not in col_map:
                col_map['nilai'] = 7
            if 'mata_pelajaran' not in col_map:
                col_map['mata_pelajaran'] = 3
            if 'semester' not in col_map:
                col_map['semester'] = 4
            if 'tahun_ajaran' not in col_map:
                col_map['tahun_ajaran'] = 5
            if 'jenis' not in col_map:
                col_map['jenis'] = 6

        # ==========================================
        # DETERMINE DATA START ROW
        # ==========================================
        # Simplified template has instruction row 2, so data starts at row 3
        # Legacy template has data starting at row 2
        start_row = 3 if is_simplified else 2
        rows = list(sheet.iter_rows(min_row=start_row, values_only=True))

        if len(rows) == 0:
            return Response({
                'success': False,
                'message': 'File kosong atau tidak ada data'
            }, status=status.HTTP_400_BAD_REQUEST)

        success_count = 0
        error_count = 0
        errors = []
        updated_count = 0

        for row_idx, row in enumerate(rows):
            # Calculate actual Excel row number for error messages
            row_num = row_idx + start_row

            try:
                # Skip empty rows and metadata rows
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                if row[0] == '[METADATA]':
                    continue

                # Get NISN (required)
                nisn_idx = col_map.get('nisn', 1)
                nisn = str(row[nisn_idx]).strip() if len(row) > nisn_idx and row[nisn_idx] else None

                if not nisn:
                    error_count += 1
                    errors.append(f'Baris {row_num}: NISN kosong')
                    continue

                # Get nilai (required)
                nilai_idx = col_map.get('nilai')
                if nilai_idx is None:
                    nilai_idx = 3 if is_simplified else 7
                nilai_raw = row[nilai_idx] if len(row) > nilai_idx else None

                if nilai_raw is None or str(nilai_raw).strip() == '':
                    # Skip rows without nilai (user didn't fill it)
                    continue

                try:
                    nilai = int(float(nilai_raw))
                except (ValueError, TypeError):
                    error_count += 1
                    errors.append(f'Baris {row_num}: Nilai "{nilai_raw}" tidak valid')
                    continue

                if nilai < 0 or nilai > 100:
                    error_count += 1
                    errors.append(f'Baris {row_num}: Nilai {nilai} harus 0-100')
                    continue

                # Get other fields (from Excel or defaults)
                # For simplified template, these come from defaults/metadata
                if is_simplified:
                    mata_pelajaran = default_mata_pelajaran
                    semester = default_semester
                    tahun_ajaran = default_tahun_ajaran
                    jenis = default_jenis
                else:
                    # Legacy template: read from Excel columns
                    mata_pelajaran_idx = col_map.get('mata_pelajaran', 3)
                    mata_pelajaran = str(row[mata_pelajaran_idx]).strip() if len(row) > mata_pelajaran_idx and row[mata_pelajaran_idx] else default_mata_pelajaran

                    semester_idx = col_map.get('semester', 4)
                    semester = str(row[semester_idx]).strip() if len(row) > semester_idx and row[semester_idx] else default_semester

                    tahun_ajaran_idx = col_map.get('tahun_ajaran', 5)
                    tahun_ajaran = str(row[tahun_ajaran_idx]).strip() if len(row) > tahun_ajaran_idx and row[tahun_ajaran_idx] else default_tahun_ajaran

                    jenis_idx = col_map.get('jenis', 6)
                    jenis = str(row[jenis_idx]).strip() if len(row) > jenis_idx and row[jenis_idx] else default_jenis

                # Validate required fields
                if not mata_pelajaran:
                    error_count += 1
                    errors.append(f'Baris {row_num}: Mata pelajaran kosong')
                    continue

                # Look up student by NISN
                try:
                    student = Student.objects.get(nisn=nisn)
                except Student.DoesNotExist:
                    error_count += 1
                    errors.append(f'Baris {row_num}: Siswa dengan NISN {nisn} tidak ditemukan')
                    continue

                # Use student's class if default_kelas not provided
                kelas = default_kelas if default_kelas else student.kelas

                # Check for existing grade (upsert)
                existing_grade = Grade.objects.filter(
                    nisn=student,
                    mata_pelajaran=mata_pelajaran,
                    semester=semester,
                    tahun_ajaran=tahun_ajaran,
                    jenis=jenis
                ).first()

                if existing_grade:
                    existing_grade.nilai = nilai
                    existing_grade.guru = guru_name
                    existing_grade.kelas = kelas
                    existing_grade.save()
                    updated_count += 1
                else:
                    Grade.objects.create(
                        nisn=student,
                        mata_pelajaran=mata_pelajaran,
                        nilai=nilai,
                        semester=semester,
                        tahun_ajaran=tahun_ajaran,
                        jenis=jenis,
                        kelas=kelas,
                        guru=guru_name
                    )
                    success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f'Baris {row_num}: {str(e)}')

        total_imported = success_count + updated_count
        kelas_info = default_kelas if default_kelas else 'sesuai data siswa'

        return Response({
            'success': True,
            'message': f'Berhasil mengimpor {total_imported} nilai untuk Kelas {kelas_info}',
            'summary': {
                'new_grades': success_count,
                'updated_grades': updated_count,
                'total_imported': total_imported,
                'errors': error_count,
                'kelas': kelas_info,
                'mata_pelajaran': default_mata_pelajaran
            },
            'errors': errors[:10] if errors else None  # Limit error list
        })

    except Exception as e:
        return Response({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsGuru | IsSuperAdmin])
def get_my_teaching_classes(request):
    """
    Get guru's assigned classes for grade input.
    Returns classes from active assignments (KBM/Diniyah only).
    """
    try:
        user = request.user

        # Get active TahunAjaran
        tahun_ajaran = TahunAjaran.objects.filter(is_active=True).first()
        if not tahun_ajaran:
            return Response({
                'success': True,
                'data': [],
                'message': 'Tidak ada tahun ajaran aktif'
            })

        # Superadmin can see all classes
        if user.role == 'superadmin':
            classes = Student.objects.values_list('kelas', flat=True).distinct()
            result = [{'kelas': k, 'mata_pelajaran': None} for k in sorted(set(c for c in classes if c))]
            return Response({
                'success': True,
                'data': result,
                'tahun_ajaran': tahun_ajaran.nama,
                'semester': tahun_ajaran.semester
            })

        # Get guru's teaching assignments (exclude piket & wali_kelas)
        assignments = Assignment.objects.filter(
            user=user,
            status='active',
            tahun_ajaran=tahun_ajaran.nama
        ).exclude(
            assignment_type__in=['piket', 'wali_kelas']
        ).values('kelas', 'mata_pelajaran').distinct()

        result = []
        seen_kelas = set()
        for a in assignments:
            if a['kelas'] and a['kelas'] not in seen_kelas:
                result.append({
                    'kelas': a['kelas'],
                    'mata_pelajaran': a['mata_pelajaran']
                })
                seen_kelas.add(a['kelas'])

        return Response({
            'success': True,
            'data': result,
            'tahun_ajaran': tahun_ajaran.nama,
            'semester': tahun_ajaran.semester
        })

    except Exception as e:
        return Response({
            'success': False,
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_students_by_class(request, kelas):
    """
    Get list of students in a class for grade input form.
    """
    try:
        students = Student.objects.filter(kelas=kelas, aktif=True).order_by('nama')

        if not students.exists():
            # Try case-insensitive
            students = Student.objects.filter(kelas__iexact=kelas, aktif=True).order_by('nama')

        data = []
        for s in students:
            data.append({
                'nisn': s.nisn,
                'nama': s.nama,
                'kelas': s.kelas
            })

        return Response({
            'success': True,
            'kelas': kelas,
            'count': len(data),
            'students': data
        })

    except Exception as e:
        return Response({
            'success': False,
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsGuru | IsSuperAdmin])
def input_batch_grades(request):
    """
    Batch input grades for a class.

    Request body:
    {
        "kelas": "XI A",
        "mata_pelajaran": "Matematika",
        "jenis": "penugasan",  // penugasan|tes_tulis|tes_lisan|portofolio|praktek|proyek|uts|uas
        "semester": "Genap",
        "tahun_ajaran": "2024/2025",
        "materi": "Bab 5 - Integral",  // optional
        "data": [
            {"nisn": "0012345678", "nilai": 85},
            {"nisn": "0012345679", "nilai": 90}
        ]
    }
    """
    try:
        data = request.data

        # Validate required fields
        kelas = data.get('kelas')
        mata_pelajaran = data.get('mata_pelajaran')
        jenis = data.get('jenis', 'penugasan')
        semester = data.get('semester', 'Ganjil')
        tahun_ajaran = data.get('tahun_ajaran', '2024/2025')
        materi = data.get('materi', '')  # New field - optional
        grades_data = data.get('data', [])

        if not kelas:
            return Response({
                'success': False,
                'message': 'Kelas harus diisi'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not mata_pelajaran:
            return Response({
                'success': False,
                'message': 'Mata pelajaran harus diisi'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not grades_data:
            return Response({
                'success': False,
                'message': 'Data nilai kosong'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get guru name
        guru_name = request.user.name if hasattr(request.user, 'name') and request.user.name else request.user.username

        success_count = 0
        updated_count = 0
        error_count = 0
        errors = []

        for item in grades_data:
            try:
                nisn = str(item.get('nisn', '')).strip()
                nilai_raw = item.get('nilai')

                # Skip empty entries
                if not nisn or nilai_raw is None or nilai_raw == '':
                    continue

                # Validate nilai
                try:
                    nilai = int(float(nilai_raw))
                except (ValueError, TypeError):
                    error_count += 1
                    errors.append(f'NISN {nisn}: Nilai tidak valid')
                    continue

                if nilai < 0 or nilai > 100:
                    error_count += 1
                    errors.append(f'NISN {nisn}: Nilai harus 0-100')
                    continue

                # Find student
                try:
                    student = Student.objects.get(nisn=nisn)
                except Student.DoesNotExist:
                    error_count += 1
                    errors.append(f'NISN {nisn}: Siswa tidak ditemukan')
                    continue

                # Check/update existing grade or create new
                existing = Grade.objects.filter(
                    nisn=student,
                    mata_pelajaran=mata_pelajaran,
                    jenis=jenis,
                    semester=semester,
                    tahun_ajaran=tahun_ajaran
                ).first()

                if existing:
                    existing.nilai = nilai
                    existing.guru = guru_name
                    if materi:  # Update materi only if provided
                        existing.materi = materi
                    existing.save()
                    updated_count += 1
                else:
                    Grade.objects.create(
                        nisn=student,
                        mata_pelajaran=mata_pelajaran,
                        nilai=nilai,
                        jenis=jenis,
                        semester=semester,
                        tahun_ajaran=tahun_ajaran,
                        kelas=kelas,
                        guru=guru_name,
                        materi=materi
                    )
                    success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f'Error: {str(e)}')

        total = success_count + updated_count

        return Response({
            'success': True,
            'message': f'Berhasil menyimpan {total} nilai',
            'summary': {
                'new_grades': success_count,
                'updated_grades': updated_count,
                'total': total,
                'errors': error_count
            },
            'errors': errors[:5] if errors else None
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mapel_list(request):
    """
    Get list of mata pelajaran for grade input dropdown.

    Logic:
    - guru/musyrif: return unique mata_pelajaran from active Assignments
    - superadmin/admin/pimpinan: return all active MasterMapel

    Response:
    {
        "success": true,
        "mapel_list": ["Matematika", "Bahasa Indonesia", ...]
    }
    """
    user = request.user

    # For guru/musyrif: get from their active assignments
    if user.role in ['guru', 'musyrif']:
        # Get active TahunAjaran
        tahun_ajaran = TahunAjaran.objects.filter(is_active=True).first()

        if not tahun_ajaran:
            return Response({
                'success': True,
                'mapel_list': [],
                'message': 'Tidak ada tahun ajaran aktif'
            })

        # Get unique mata_pelajaran from active assignments (excluding piket, wali_kelas)
        assignments = Assignment.objects.filter(
            user=user,
            status='active',
            tahun_ajaran=tahun_ajaran.nama
        ).exclude(
            assignment_type__in=['piket', 'wali_kelas']
        ).values_list('mata_pelajaran', flat=True).distinct()

        # Filter out None/empty values and sort
        mapel_list = sorted([mp for mp in assignments if mp])

        return Response({
            'success': True,
            'mapel_list': mapel_list
        })

    # For superadmin/admin/pimpinan: get all active MasterMapel
    elif user.role in ['superadmin', 'admin', 'pimpinan']:
        master_mapel = MasterMapel.objects.filter(
            is_active=True
        ).values_list('nama', flat=True).distinct()

        mapel_list = sorted(list(set(master_mapel)))

        return Response({
            'success': True,
            'mapel_list': mapel_list
        })

    # Other roles: return empty or limited list
    else:
        return Response({
            'success': True,
            'mapel_list': [],
            'message': 'Role tidak memiliki akses ke daftar mata pelajaran'
        })
