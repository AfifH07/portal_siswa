"""
Robust XLSX Parser & Template Generator for Student Data Import
===============================================================
Features:
- Openpyxl-based Excel reading (no CSV)
- Smart header detection (finds NISN/NAMA SISWA row)
- Column sanitization (string NISN, strip whitespace, ISO dates)
- Styled template generation
- User-friendly error messages
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime
from typing import Tuple, List, Dict, Any, Optional
import re


# ============================================================
# COLUMN MAPPINGS - Flexible header recognition
# ============================================================

COLUMN_MAPPINGS = {
    'nisn': ['nisn', 'nomor induk siswa nasional', 'no_induk_nasional', 'nomor induk nasional'],
    'nis': ['nis', 'no_induk', 'nomor_induk', 'no induk', 'nomor induk siswa',
            'nomor induk', 'no. induk', 'nomer induk', 'nomor induk lokal'],
    'nama': ['nama', 'nama_lengkap', 'nama lengkap', 'nama_siswa', 'nama siswa',
             'fullname', 'name', 'nama santri', 'nama_santri'],
    'kelas': ['kelas', 'class', 'tingkat', 'grade', 'rombel', 'rombongan belajar'],
    'program': ['program', 'jurusan', 'peminatan', 'major', 'konsentrasi'],
    'jenis_kelamin': ['jenis_kelamin', 'jenis kelamin', 'gender', 'kelamin', 'jk',
                      'l/p', 'laki/perempuan', 'sex'],
    'status': ['status', 'aktif', 'active', 'status_aktif', 'status siswa', 'keterangan'],
    'email': ['email', 'e-mail', 'email_siswa', 'alamat email'],
    'phone': ['phone', 'hp', 'no_hp', 'telepon', 'no_telepon', 'handphone', 'no hp',
              'nomor hp', 'nomor telepon', 'telp'],
    'wali_nama': ['wali_nama', 'nama_wali', 'wali', 'guardian', 'orangtua', 'orang_tua',
                  'nama wali', 'nama orang tua', 'nama ayah', 'nama ibu'],
    'wali_phone': ['wali_phone', 'hp_wali', 'no_hp_wali', 'telepon_wali', 'hp wali',
                   'no hp wali', 'telp wali', 'nomor wali'],
    'target_hafalan': ['target_hafalan', 'target hafalan', 'target_juz', 'target juz',
                       'target', 'juz target'],
    'current_hafalan': ['current_hafalan', 'hafalan_sekarang', 'hafalan sekarang',
                        'juz_sekarang', 'current hafalan', 'hafalan', 'juz hafalan',
                        'capaian hafalan'],
    'target_nilai': ['target_nilai', 'target nilai', 'kkm', 'nilai target'],
    'tanggal_masuk': ['tanggal_masuk', 'tanggal masuk', 'tgl_masuk', 'tgl masuk',
                      'entry_date', 'admission_date', 'tanggal daftar', 'tgl daftar'],
    'catatan': ['catatan', 'keterangan', 'note', 'notes', 'catatan siswa', 'remark'],
}

# Keywords to identify header row
HEADER_KEYWORDS = ['nisn', 'nis', 'nama', 'nama siswa', 'nama santri', 'nama lengkap']


# ============================================================
# SANITIZATION FUNCTIONS
# ============================================================

def sanitize_string(value: Any) -> str:
    """
    Sanitize string value:
    - Remove leading/trailing whitespace
    - Remove leading single quotes (')
    - Convert to string
    """
    if pd.isna(value) or value is None:
        return ''

    # Convert to string
    str_value = str(value).strip()

    # Remove leading single quote (often added by Excel to force text format)
    if str_value.startswith("'"):
        str_value = str_value[1:]

    # Remove trailing .0 for numbers that were converted to float
    if str_value.endswith('.0') and str_value[:-2].replace('-', '').isdigit():
        str_value = str_value[:-2]

    return str_value.strip()


def sanitize_nisn(value: Any) -> str:
    """
    Specifically sanitize NISN/NIS values:
    - Force string type
    - Remove leading quotes
    - Preserve leading zeros
    - Remove non-alphanumeric characters (except dash)
    """
    if pd.isna(value) or value is None:
        return ''

    # Convert to string, handling float formatting
    if isinstance(value, (int, float)):
        # Preserve as integer string (no decimals)
        str_value = str(int(value))
    else:
        str_value = str(value).strip()

    # Remove leading quote
    if str_value.startswith("'"):
        str_value = str_value[1:]

    # Remove trailing .0
    if str_value.endswith('.0'):
        str_value = str_value[:-2]

    # Remove extra whitespace
    str_value = str_value.strip()

    # Remove any non-alphanumeric characters except dash
    str_value = re.sub(r'[^\w-]', '', str_value)

    return str_value


def sanitize_name(value: Any) -> str:
    """
    Sanitize name:
    - Strip whitespace
    - Capitalize properly
    - Remove extra spaces
    """
    if pd.isna(value) or value is None:
        return ''

    str_value = str(value).strip()

    # Remove multiple spaces
    str_value = re.sub(r'\s+', ' ', str_value)

    return str_value.strip()


def parse_status(value: Any) -> bool:
    """Parse status value to boolean."""
    if pd.isna(value) or value is None:
        return True  # Default to active

    str_value = str(value).lower().strip()

    # Check for inactive indicators
    inactive_values = ['tidak aktif', 'tidak', 'non-aktif', 'nonaktif', 'inactive',
                       'false', '0', 'no', 'n', 'alumni', 'lulus', 'keluar', 'pindah',
                       'drop out', 'do', 'berhenti']

    return str_value not in inactive_values


def parse_jenis_kelamin(value: Any) -> Optional[str]:
    """
    Parse jenis kelamin value to 'L' or 'P'.
    Returns None if not recognized.
    """
    if pd.isna(value) or value is None:
        return None

    str_value = str(value).upper().strip()

    # Check for male indicators
    male_values = ['L', 'LAKI-LAKI', 'LAKI', 'MALE', 'M', 'PRIA', 'COWOK', '1']
    if str_value in male_values:
        return 'L'

    # Check for female indicators
    female_values = ['P', 'PEREMPUAN', 'FEMALE', 'F', 'WANITA', 'CEWEK', '2']
    if str_value in female_values:
        return 'P'

    return None


def parse_integer(value: Any, default: int = 0) -> int:
    """Parse value to integer."""
    if pd.isna(value) or value is None:
        return default

    try:
        str_value = sanitize_string(value)
        if not str_value:
            return default
        return int(float(str_value))
    except (ValueError, TypeError):
        return default


def parse_date_to_iso(value: Any) -> Optional[str]:
    """
    Parse date value to ISO format (YYYY-MM-DD).
    Handles various date formats.
    """
    if pd.isna(value) or value is None:
        return None

    # If already a datetime
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime('%Y-%m-%d')

    str_value = str(value).strip()
    if not str_value:
        return None

    # Remove time component if present
    str_value = str_value.split(' ')[0]

    # Try common date formats
    date_formats = [
        '%Y-%m-%d',      # 2024-01-15
        '%d-%m-%Y',      # 15-01-2024
        '%d/%m/%Y',      # 15/01/2024
        '%Y/%m/%d',      # 2024/01/15
        '%d-%m-%y',      # 15-01-24
        '%d/%m/%y',      # 15/01/24
        '%Y%m%d',        # 20240115
        '%d %B %Y',      # 15 January 2024
        '%d %b %Y',      # 15 Jan 2024
    ]

    for fmt in date_formats:
        try:
            parsed = datetime.strptime(str_value, fmt)
            return parsed.strftime('%Y-%m-%d')
        except ValueError:
            continue

    return None


# ============================================================
# SMART HEADER DETECTION
# ============================================================

def find_header_row_openpyxl(ws, max_rows: int = 30) -> int:
    """
    Find the row containing column headers by searching for known keywords.
    Uses openpyxl worksheet directly for better accuracy.

    Returns the row number (1-indexed) where headers are found.
    """
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_values.append(str(cell_value).lower().strip())
            else:
                row_values.append('')

        # Count how many known header keywords are found in this row
        matches = sum(1 for val in row_values if val in HEADER_KEYWORDS)

        # If we find at least 2 matching headers (NISN and NAMA), this is the header row
        if matches >= 2:
            return row_idx

    # Default to first row if no header row found
    return 1


def map_columns_from_row(ws, header_row: int) -> Dict[str, int]:
    """
    Map column indices from header row.
    Returns dict of {standardized_name: column_index (1-indexed)}
    """
    column_map = {}

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if not cell_value:
            continue

        header_lower = str(cell_value).lower().strip()

        # Try to match with known column names
        for std_name, variations in COLUMN_MAPPINGS.items():
            if std_name in column_map:
                continue  # Already mapped

            for var in variations:
                if var.lower() == header_lower:
                    column_map[std_name] = col_idx
                    break

    return column_map


# ============================================================
# EXCEL READING & PARSING
# ============================================================

def read_excel_robust(file_content: bytes) -> Tuple[Any, str, int, Dict[str, int]]:
    """
    Read Excel file with smart header detection using openpyxl.

    Returns: (worksheet, error_message, header_row, column_map)
    """
    try:
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active

        if ws.max_row < 2:
            return None, "File Excel kosong atau hanya memiliki satu baris", 0, {}

        # Find header row
        header_row = find_header_row_openpyxl(ws)

        # Map columns
        column_map = map_columns_from_row(ws, header_row)

        # Validate required columns
        if 'nisn' not in column_map:
            return None, "Kolom NISN tidak ditemukan. Pastikan ada kolom dengan header 'NISN' atau 'NIS'", 0, {}

        if 'nama' not in column_map:
            return None, "Kolom NAMA tidak ditemukan. Pastikan ada kolom dengan header 'Nama' atau 'Nama Siswa'", 0, {}

        return ws, '', header_row, column_map

    except Exception as e:
        return None, f"Gagal membaca file Excel: {str(e)}", 0, {}


def parse_student_rows(ws, header_row: int, column_map: Dict[str, int]) -> Tuple[List[Dict], List[Dict]]:
    """
    Parse worksheet rows into student records.
    Returns (valid_records, error_records)
    """
    from .models import Student

    valid_records = []
    error_records = []

    # Get existing NISNs for duplicate check
    existing_nisns = set(Student.objects.values_list('nisn', flat=True))

    # Start from row after header
    data_start_row = header_row + 2

    for row_idx in range(data_start_row, ws.max_row + 1):
        # Get NISN first to check if row has data
        nisn_col = column_map.get('nisn')
        nisn_value = ws.cell(row=row_idx, column=nisn_col).value if nisn_col else None

        # Skip empty rows
        if not nisn_value or str(nisn_value).strip() == '':
            continue
        
        nisn = sanitize_nisn(nisn_value)

        # Get nama
        nama_col = column_map.get('nama')
        nama_value = ws.cell(row=row_idx, column=nama_col).value if nama_col else None
        nama = sanitize_name(nama_value)

        # Validate NISN
        if not nisn:
            error_records.append({
                'row': row_idx,
                'nisn': '',
                'nama': nama,
                'message': f"Baris {row_idx}: NISN kosong atau tidak valid"
            })
            continue

        # Validate nama
        if not nama:
            error_records.append({
                'row': row_idx,
                'nisn': nisn,
                'nama': '',
                'message': f"Baris {row_idx}: Nama siswa kosong"
            })
            continue

        # Check for duplicate NISN
        if nisn in existing_nisns:
            error_records.append({
                'row': row_idx,
                'nisn': nisn,
                'nama': nama,
                'message': f"Gagal mengimpor: '{nama}' di baris {row_idx} memiliki NISN yang sudah terdaftar ({nisn})"
            })
            continue

        # Helper function to get cell value
        def get_cell(col_name):
            col_idx = column_map.get(col_name)
            if col_idx:
                return ws.cell(row=row_idx, column=col_idx).value
            return None

        # Build student record
        try:
            # Parse NIS (local student number)
            nis_value = sanitize_nisn(get_cell('nis'))
            nis = nis_value if nis_value else None

            student = {
                'nisn': nisn,
                'nis': nis,
                'nama': nama,
                'jenis_kelamin': parse_jenis_kelamin(get_cell('jenis_kelamin')),
                'kelas': sanitize_string(get_cell('kelas')) or None,
                'program': sanitize_string(get_cell('program')) or 'Reguler',
                'email': sanitize_string(get_cell('email')) or None,
                'phone': sanitize_string(get_cell('phone')) or None,
                'wali_nama': sanitize_string(get_cell('wali_nama')) or None,
                'wali_phone': sanitize_string(get_cell('wali_phone')) or None,
                'target_hafalan': parse_integer(get_cell('target_hafalan'), 0),
                'current_hafalan': parse_integer(get_cell('current_hafalan'), 0),
                'target_nilai': parse_integer(get_cell('target_nilai'), 75),
                'tanggal_masuk': parse_date_to_iso(get_cell('tanggal_masuk')),
                'aktif': parse_status(get_cell('status')),
                'catatan': sanitize_string(get_cell('catatan')) or '',
                '_row': row_idx
            }

            valid_records.append(student)
            existing_nisns.add(nisn)  # Add to set to catch duplicates within file

        except Exception as e:
            error_records.append({
                'row': row_idx,
                'nisn': nisn,
                'nama': nama,
                'message': f"Baris {row_idx}: Error memproses data - {str(e)}"
            })

    return valid_records, error_records


# ============================================================
# ERROR REPORT GENERATION
# ============================================================

def generate_error_report_xlsx(error_records: List[Dict]) -> bytes:
    """
    Generate styled Excel error report.
    Returns Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Error Import"

    # Styles
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    error_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Baris', 'NISN', 'Nama', 'Pesan Error']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    for row_idx, err in enumerate(error_records, 2):
        ws.cell(row=row_idx, column=1, value=err.get('row', '-')).border = border
        ws.cell(row=row_idx, column=2, value=err.get('nisn', '')).border = border
        ws.cell(row=row_idx, column=3, value=err.get('nama', '')).border = border

        error_cell = ws.cell(row=row_idx, column=4, value=err.get('message', ''))
        error_cell.fill = error_fill
        error_cell.border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 60

    # Freeze header row
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# TEMPLATE GENERATION
# ============================================================

def generate_import_template() -> bytes:
    """
    Generate styled Excel import template with:
    - Colored headers (Emerald Green)
    - Sample data row
    - NISN column formatted as Text
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Siswa"

    # Baron Emerald color scheme
    emerald_fill = PatternFill(start_color="178560", end_color="178560", fill_type="solid")
    emerald_light_fill = PatternFill(start_color="D6F5EC", end_color="D6F5EC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    sample_font = Font(italic=True, color="666666", size=10)
    border = Border(
        left=Side(style='thin', color='AEEBD8'),
        right=Side(style='thin', color='AEEBD8'),
        top=Side(style='thin', color='AEEBD8'),
        bottom=Side(style='thin', color='AEEBD8')
    )

    # Column definitions with widths
    columns = [
        ('NISN', 15, 'Wajib - Nomor Induk Siswa Nasional'),
        ('NIS', 12, 'Wajib - Nomor Induk Siswa lokal (7 digit)'),
        ('Nama', 30, 'Wajib - Nama lengkap siswa'),
        ('Kelas', 10, 'Contoh: X A, XI B, XII C'),
        ('Program', 15, 'Reguler / Tahfidz / Khusus'),
        ('Jenis Kelamin', 14, 'L (Laki-laki) / P (Perempuan)'),
        ('Status', 12, 'Aktif / Tidak Aktif'),
        ('Email', 25, 'Alamat email siswa'),
        ('No HP', 15, 'Nomor telepon siswa'),
        ('Nama Wali', 25, 'Nama orang tua/wali'),
        ('No HP Wali', 15, 'Nomor telepon wali'),
        ('Target Hafalan', 14, 'Target juz (angka)'),
        ('Hafalan Sekarang', 16, 'Capaian juz saat ini'),
        ('Target Nilai', 12, 'KKM (default: 75)'),
        ('Tanggal Masuk', 14, 'Format: DD/MM/YYYY'),
        ('Catatan', 30, 'Catatan tambahan (opsional)'),
    ]

    # Sample data
    sample_row = [
        "'0069028700",  # Leading quote to force text
        "'0000664",     # Leading quote to force text
        "Contoh Nama Siswa",
        "XII A",
        "Tahfidz",
        "L",
        "Aktif",
        "",
        "",
        "",
        "",
        "30",
        "15",
        "75",
        "01/01/2024",
        "",
    ]

    # Write headers (Row 1)
    for col_idx, (header, width, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = emerald_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    # Write hint row (Row 2) - Column descriptions
    for col_idx, (_, _, hint) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.fill = emerald_light_fill
        cell.font = Font(italic=True, color="3D6B57", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Write sample data (Row 3)
    for col_idx, value in enumerate(sample_row, 1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.font = sample_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[3].height = 20

    # Freeze header rows
    ws.freeze_panes = 'A4'

    # Set NISN and NIS columns to Text format
    for row in range(3, 1000):  # Format many rows for future data
        ws.cell(row=row, column=1).number_format = '@'  # NISN - Text format
        ws.cell(row=row, column=2).number_format = '@'  # NIS - Text format

    # Add instruction sheet
    ws_info = wb.create_sheet(title="Petunjuk Pengisian")

    instructions = [
        ("PETUNJUK PENGISIAN DATA SISWA", None),
        ("", None),
        ("1. Kolom NISN, NIS, dan Nama WAJIB diisi", None),
        ("2. Untuk NISN/NIS yang diawali angka 0, tambahkan tanda petik (') di depan", None),
        ("   Contoh NISN: '0069028700", None),
        ("   Contoh NIS: '0000664 (7 digit)", None),
        ("3. Jenis Kelamin diisi dengan: L (Laki-laki) atau P (Perempuan)", None),
        ("4. Format tanggal: DD/MM/YYYY (contoh: 01/01/2024)", None),
        ("5. Status diisi dengan: Aktif atau Tidak Aktif", None),
        ("6. Program diisi dengan: Reguler, Tahfidz, atau Khusus", None),
        ("7. Hapus baris contoh (baris 3) sebelum mengimport data asli", None),
        ("8. Baris petunjuk (baris 2) boleh dihapus atau dibiarkan", None),
        ("9. Kolom Catatan bersifat opsional", None),
        ("", None),
        ("KELAS YANG TERSEDIA:", None),
        ("Kelas X: X A, X B, X C, X D", None),
        ("Kelas XI: XI A, XI B, XI C, XI D", None),
        ("Kelas XII: XII A, XII B, XII C, XII D", None),
    ]

    for row_idx, (text, _) in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="178560")
        elif text.startswith("KELAS"):
            cell.font = Font(bold=True, size=11, color="178560")
        else:
            cell.font = Font(size=10)

    ws_info.column_dimensions['A'].width = 70

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# MAIN IMPORT FUNCTION
# ============================================================

def import_students(file_content: bytes, filename: str) -> Dict[str, Any]:
    """
    Main function to import students from Excel file.

    Returns:
    {
        'success': bool,
        'created': int,
        'updated': int,
        'skipped': int,
        'errors': list of user-friendly messages,
        'error_file': bytes or None
    }
    """
    from .models import Student

    # Validate file extension
    if not filename.lower().endswith(('.xlsx', '.xls')):
        return {
            'success': False,
            'error': 'Format file tidak didukung. Gunakan file Excel (.xlsx)',
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': ['Harap gunakan file Excel (.xlsx) yang didownload dari template'],
            'error_file': None
        }

    # Read Excel file
    ws, read_error, header_row, column_map = read_excel_robust(file_content)

    if read_error:
        return {
            'success': False,
            'error': read_error,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': [read_error],
            'error_file': None
        }

    # Parse student data
    valid_records, error_records = parse_student_rows(ws, header_row, column_map)

    if len(valid_records) == 0 and len(error_records) == 0:
        return {
            'success': False,
            'error': 'Tidak ada data siswa yang ditemukan dalam file',
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': ['File tidak memiliki data siswa yang valid. Pastikan data dimulai setelah baris header.'],
            'error_file': None
        }

    created_count = 0
    updated_count = 0

    # Process valid records
    for student_data in valid_records:
        row_num = student_data.pop('_row', None)
        nisn = student_data['nisn']
        nama = student_data['nama']

        try:
            # Check if student exists (for update scenario)
            existing = Student.objects.filter(nisn=nisn).first()

            if existing:
                # Update existing student
                for key, value in student_data.items():
                    if value is not None and value != '':
                        setattr(existing, key, value)
                existing.save()
                updated_count += 1
            else:
                # Create new student
                Student.objects.create(**student_data)
                created_count += 1

        except Exception as e:
            error_msg = str(e)
            if 'unique' in error_msg.lower() or 'duplicate' in error_msg.lower():
                error_records.append({
                    'row': row_num,
                    'nisn': nisn,
                    'nama': nama,
                    'message': f"Gagal mengimpor: '{nama}' di baris {row_num} memiliki NISN yang sudah terdaftar"
                })
            else:
                error_records.append({
                    'row': row_num,
                    'nisn': nisn,
                    'nama': nama,
                    'message': f"Gagal menyimpan data '{nama}' di baris {row_num}: {error_msg}"
                })

    # Prepare user-friendly error messages
    user_errors = [err['message'] for err in error_records[:15]]  # Limit to 15 messages

    if len(error_records) > 15:
        user_errors.append(f"... dan {len(error_records) - 15} error lainnya. Download laporan error untuk detail lengkap.")

    # Generate error file if there are errors
    error_file = None
    if error_records:
        try:
            error_file = generate_error_report_xlsx(error_records)
        except Exception:
            pass

    total_processed = created_count + updated_count
    success = total_processed > 0

    return {
        'success': success,
        'created': created_count,
        'updated': updated_count,
        'skipped': len(error_records),
        'errors': user_errors,
        'error_file': error_file,
        'message': f"Berhasil mengimpor {total_processed} siswa ({created_count} baru, {updated_count} diperbarui)"
                   if success else "Tidak ada data yang berhasil diimpor"
    }
