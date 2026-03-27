"""
Script untuk mengimport data SISWA dan NILAI dari file Excel LEGER ke database.
Jalankan: python import_siswa_dan_nilai.py

Format Excel LEGER:
- Row 3: Kelas
- Row 4-7: Header
- Row 8+: Data siswa (NO, NAMA, NISN, NIS, nilai...)
"""

import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend_django.settings')
django.setup()

import openpyxl
from apps.grades.models import Grade
from apps.students.models import Student


# Mapping semester Excel label to database format
SEMESTER_MAPPING = {
    'smt1': ('Ganjil', '2022/2023'),   # Kelas X Semester 1
    'smt2': ('Genap', '2022/2023'),    # Kelas X Semester 2
    'smt3': ('Ganjil', '2023/2024'),   # Kelas XI Semester 1
    'smt4': ('Genap', '2023/2024'),    # Kelas XI Semester 2
    'smt5': ('Ganjil', '2024/2025'),   # Kelas XII Semester 1
    'smt6': ('Genap', '2024/2025'),    # Kelas XII Semester 2
}


def parse_leger_structure_all_semesters(sheet):
    """Parse struktur file LEGER untuk semua semester."""
    nisn_col = 3  # Column C
    nama_col = 2  # Column B

    # mapel_semester_structure: list of (col_idx, mapel_name, semester_label)
    mapel_semester_structure = []

    row5_values = [sheet.cell(row=5, column=c).value for c in range(1, sheet.max_column + 1)]
    row7_values = [sheet.cell(row=7, column=c).value for c in range(1, sheet.max_column + 1)]

    current_mapel = None

    for col_idx in range(5, len(row5_values) + 1):
        mapel_val = row5_values[col_idx - 1] if col_idx - 1 < len(row5_values) else None
        if mapel_val and str(mapel_val).strip():
            current_mapel = str(mapel_val).strip()

        smt_val = row7_values[col_idx - 1] if col_idx - 1 < len(row7_values) else None
        if smt_val and 'smt' in str(smt_val).strip().lower():
            smt_key = str(smt_val).strip().lower()
            if current_mapel and smt_key in SEMESTER_MAPPING:
                mapel_semester_structure.append((col_idx, current_mapel, smt_key))

    return nisn_col, nama_col, mapel_semester_structure


def import_siswa_dari_leger(filepath, kelas):
    """Import data siswa dari file LEGER."""
    print(f"\n{'='*60}")
    print(f"Importing SISWA dari: {os.path.basename(filepath)}")
    print(f"Kelas: {kelas}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    nisn_col = 3
    nama_col = 2
    nis_col = 4

    created_count = 0
    exists_count = 0

    for row_idx in range(8, sheet.max_row + 1):
        nisn_value = sheet.cell(row=row_idx, column=nisn_col).value
        nama_value = sheet.cell(row=row_idx, column=nama_col).value
        nis_value = sheet.cell(row=row_idx, column=nis_col).value

        if not nisn_value or not nama_value:
            continue

        nisn_str = str(nisn_value).strip()
        nama_str = str(nama_value).strip()
        nis_str = str(nis_value).strip() if nis_value else ''

        if not nisn_str.isdigit() or len(nisn_str) < 5:
            continue

        # Cek atau create student
        student, created = Student.objects.get_or_create(
            nisn=nisn_str,
            defaults={
                'nama': nama_str,
                'kelas': kelas,
                'program': 'Reguler',
                'aktif': True
            }
        )

        if created:
            created_count += 1
            print(f"  + Created: {nisn_str} - {nama_str}")
        else:
            exists_count += 1
            # Update kelas jika berbeda
            if student.kelas != kelas:
                student.kelas = kelas
                student.save()

    wb.close()

    print(f"\nHasil Import Siswa untuk {kelas}:")
    print(f"  Created: {created_count}")
    print(f"  Already exists: {exists_count}")

    return created_count


def import_grades_all_semesters(filepath, kelas, guru='Admin'):
    """Import data nilai dari SEMUA semester yang ada di file LEGER Excel."""
    print(f"\n{'='*60}")
    print(f"Importing NILAI dari: {os.path.basename(filepath)}")
    print(f"Kelas: {kelas}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    nisn_col, nama_col, mapel_semester_structure = parse_leger_structure_all_semesters(sheet)

    # Count per semester
    semester_mapel_count = {}
    for _, mapel, smt in mapel_semester_structure:
        if smt not in semester_mapel_count:
            semester_mapel_count[smt] = set()
        semester_mapel_count[smt].add(mapel)

    print(f"\n  Struktur terdeteksi:")
    for smt in sorted(semester_mapel_count.keys()):
        print(f"    {smt}: {len(semester_mapel_count[smt])} mata pelajaran")

    success_count = 0
    error_count = 0
    semester_success = {}

    for row_idx in range(8, sheet.max_row + 1):
        nisn_value = sheet.cell(row=row_idx, column=nisn_col).value

        if not nisn_value:
            continue

        nisn_str = str(nisn_value).strip()

        if not nisn_str.isdigit() or len(nisn_str) < 5:
            continue

        try:
            student = Student.objects.get(nisn=nisn_str)
        except Student.DoesNotExist:
            error_count += 1
            continue

        for col_idx, mapel_name, smt_key in mapel_semester_structure:
            nilai_value = sheet.cell(row=row_idx, column=col_idx).value

            if nilai_value is None or str(nilai_value).strip() == '':
                continue

            try:
                nilai = int(float(nilai_value))

                if nilai < 0 or nilai > 100:
                    continue

                semester, tahun_ajaran = SEMESTER_MAPPING[smt_key]

                existing = Grade.objects.filter(
                    nisn=student,
                    mata_pelajaran=mapel_name,
                    semester=semester,
                    tahun_ajaran=tahun_ajaran,
                    jenis='UAS'
                ).first()

                if existing:
                    existing.nilai = nilai
                    existing.kelas = kelas
                    existing.guru = guru
                    existing.save()
                else:
                    Grade.objects.create(
                        nisn=student,
                        mata_pelajaran=mapel_name,
                        nilai=nilai,
                        semester=semester,
                        tahun_ajaran=tahun_ajaran,
                        jenis='UAS',
                        kelas=kelas,
                        guru=guru
                    )

                success_count += 1
                semester_success[smt_key] = semester_success.get(smt_key, 0) + 1

            except (ValueError, TypeError):
                error_count += 1

    wb.close()

    print(f"\nHasil Import Nilai untuk {kelas}:")
    print(f"  Total Berhasil: {success_count}")
    for smt in sorted(semester_success.keys()):
        s, t = SEMESTER_MAPPING[smt]
        print(f"    {smt} ({s} {t}): {semester_success[smt]}")
    print(f"  Error: {error_count}")

    return success_count


def main():
    data_folder = r'C:\Users\Afif H\.vscode\belajaroiii\semester 6\portal-siswa\data nilai'

    print("="*80)
    print("IMPORT SISWA DAN NILAI DARI FILE LEGER EXCEL")
    print("(Semua semester: Smt1-Smt6)")
    print("="*80)

    files = [
        ('f_legersemua_XII-A (2).xlsx', 'XII A'),
        ('f_legersemua_XII-B (2).xlsx', 'XII B'),
        ('f_legersemua_XII-C (2).xlsx', 'XII C'),
    ]

    # Step 1: Import Siswa
    print("\n" + "="*80)
    print("STEP 1: IMPORT DATA SISWA")
    print("="*80)

    total_siswa_created = 0
    for filename, kelas in files:
        filepath = os.path.join(data_folder, filename)
        if os.path.exists(filepath):
            count = import_siswa_dari_leger(filepath, kelas)
            total_siswa_created += count

    print(f"\nTotal Siswa Baru: {total_siswa_created}")

    # Step 2: Import Nilai (ALL semesters)
    print("\n" + "="*80)
    print("STEP 2: IMPORT DATA NILAI (SEMUA SEMESTER)")
    print("="*80)

    total_nilai = 0
    for filename, kelas in files:
        filepath = os.path.join(data_folder, filename)
        if os.path.exists(filepath):
            count = import_grades_all_semesters(
                filepath,
                kelas=kelas,
                guru='Admin'
            )
            total_nilai += count

    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total Siswa Baru: {total_siswa_created}")
    print(f"Total Nilai Diimport: {total_nilai}")

    # Verify
    print("\n=== VERIFIKASI DATABASE ===")
    print(f"Total Students: {Student.objects.count()}")
    print(f"Total Grades: {Grade.objects.count()}")

    from django.db.models import Count
    kelas_count = Student.objects.values('kelas').annotate(count=Count('id')).order_by('kelas')
    for k in kelas_count:
        print(f"  {k['kelas']}: {k['count']} siswa")

    # Grade breakdown by semester
    print("\nNilai per Semester/Tahun Ajaran:")
    grade_count = Grade.objects.values('semester', 'tahun_ajaran').annotate(count=Count('id')).order_by('tahun_ajaran', 'semester')
    for g in grade_count:
        print(f"  {g['semester']} {g['tahun_ajaran']}: {g['count']} nilai")


if __name__ == '__main__':
    main()
